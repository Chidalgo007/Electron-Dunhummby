from openpyxl import load_workbook
import datetime
import sys

def updateDateToMonday(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        today = datetime.date.today()
        # weekday(): Monday=0, Sunday=6
        monday = today - datetime.timedelta(days=today.weekday())

        ws['A2'].value = monday
        wb.save(file_path)
        return True
    except:
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(1)

    file_path = sys.argv[1]
    success = updateDateToMonday(file_path)
    sys.exit(0 if success else 1)
